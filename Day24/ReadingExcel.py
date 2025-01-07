import openpyxl

file=r"C:\Users\Prashant\Downloads\Toyota.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Data1"]
# sheet=workbook.active

rows=sheet.max_row
columns=sheet.max_column

for r in range(1,4):
    for c in range(1,4):
        print(sheet.cell(r,c).value,end="          ")
    print()