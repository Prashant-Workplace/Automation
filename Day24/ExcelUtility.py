import openpyxl


# define row count in sheet
def rowcount(filename, sheetname):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return sheet.max_row


# rowc=rowcount(filename,sheetname)
# print(row)

# define column count in sheet
def columncount(filename, sheetname):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return sheet.max_column


# colc=columncount(filename,sheetname)
# print(col)

# define read data in sheet
def readdata(filename, sheetname, row, column):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return sheet.cell(row, column).value


# read=readdata(filename,sheetname,row,column)
# print(read)

# define write data in sheet
def writedata(filename, sheetname, row, column, write):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    sheet.cell(row, column).value = write
    workbook.save(filename)
