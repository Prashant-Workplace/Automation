import openpyxl

file = "C:\\Users\\Prashant\\Downloads\\Toyota.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Data"]

# rows=sheet.max_row
# columns=sheet.max_column

# for r in range(1, 11):
#     for c in range(1, 16386):
#         sheet.cell(r, c).value = "welcome"

sheet.cell(1,1).value="name1"
sheet.cell(1,2).value="role1"
sheet.cell(1,3).value="salary1"

sheet.cell(2,1).value="name2"
sheet.cell(2,2).value="role2"
sheet.cell(2,3).value="salary2"

sheet.cell(3,1).value="name3"
sheet.cell(3,2).value="role3"
sheet.cell(3,3).value="salary3"

workbook.save(file)
